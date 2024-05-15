import re
from fastapi import HTTPException

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from models.base_table_name import TableBaseModel

from services.catalogs.clinic import ClinicService
from services.medical_record.dispensary import DispensaryService
from services.medical_record.medical_record import MedicalRecordService
from services.medical_record.parent import ParentService
from services.medical_record.visit_specialist_control import VisitSpecialistControlService
from services.medical_record.allergy import AllergyService
from services.medical_record.extra_class import ExtraClassService
from services.medical_record.past_illness import PastIllnessService
from services.medical_record.hospitalization import HospitalizationService
from services.medical_record.spa_treatment import SpaTreatmentService
from services.medical_record.medical_certificate import MedicalCertificateService
from services.medical_record.deworming import DewormingService
from services.medical_record.oral_sanation import OralSanationService
from services.medical_record.prevaccination_checkup import PrevaccinationCheckupService
from services.medical_record.vaccination import VaccinationService
from services.medical_record.gg_injection import GammaGlobulinInjectionService
from services.medical_record.mantoux_test import MantouxTestService
from services.medical_record.tub_vac import TuberculosisVaccinationService
from services.medical_record.medical_examination import MedicalExaminationService
from services.medical_record.ongoing_medical_supervision import OngoingMedicalSupervisionService
from services.medical_record.screening import ScreeningService
from services.catalogs.vac_name import VacNameService

from models.catalogs.clinic import ClinicCreate
from models.medical_record.child import ChildCreate
from models.medical_record.dispensary import DispensaryCreate
from models.medical_record.parent import ParentCreate, ParentType
from models.medical_record.visit_specialist_control import VisitSpecialistControlCreate
from models.medical_record.allergy import AllergyCreate
from models.medical_record.deworming import DewormingCreate
from models.medical_record.extra_class import ExtraClassCreate
from models.medical_record.gg_injection import GammaGlobulinInjectionCreate
from models.medical_record.hospitalization import HospitalizationCreate
from models.medical_record.mantoux_test import MantouxTestCreate
from models.medical_record.medical_certificate import MedicalCertificateCreate
from models.medical_record.medical_examination import MedicalExaminationCreate
from models.medical_record.ongoing_medical_supervision import OngoingMedicalSupervisionCreate
from models.medical_record.oral_sanation import OralSanationCreate
from models.medical_record.past_illness import PastIllnessCreate
from models.medical_record.prevaccination_checkup import PrevaccinationCheckupCreate
from models.medical_record.screening import ScreeningCreate
from models.medical_record.spa_treatment import SpaTreatmentCreate
from models.medical_record.tub_vac import TuberculosisVaccinationCreate
from models.catalogs.vac_name import VacNameCreate
from models.medical_record.vaccination import VaccinationCreate



def create_obj_generator_from_ws(ws: Worksheet, tab_name: str):
    enum_table = {
        "TableBaseModel": TableBaseModel
    }
    obj_gen = ws.values
    headers = next(obj_gen)
    for data in obj_gen:
        data_as_dict = dict(zip(headers, data))
        obj = enum_table["TableBaseModel"][tab_name].value(**data_as_dict)
        yield obj

def camel_to_snake(camel_case_string):
    parts = re.findall(r'[A-Z][a-z0-9]*', camel_case_string)
    return '_'.join(part.lower() for part in parts)



def import_from_xlsx(filename: str,
                     kindergarten_num: int,
                     medcard_service: MedicalRecordService,
                     clinic_service: ClinicService,
                     parent_service: ParentService,
                     dispensary_service: DispensaryService,
                     visit_s_control_service: VisitSpecialistControlService,
                     allergy_service: AllergyService,
                     extra_class_service: ExtraClassService,
                     past_illness_service: PastIllnessService,
                     hospitalization_service: HospitalizationService,
                     spa_treatment_service: SpaTreatmentService,
                     medical_certificate_service: MedicalCertificateService,
                     deworming_service: DewormingService,
                     oral_sanation_service: OralSanationService,
                     prevaccination_checkup_service: PrevaccinationCheckupService,
                     vaccination_service: VaccinationService,
                     gg_injection_service: GammaGlobulinInjectionService,
                     mantoux_test_service: MantouxTestService,
                     tub_vac_service: TuberculosisVaccinationService,
                     medical_examination_service: MedicalExaminationService,
                     ongoing_medical_supervision_service: OngoingMedicalSupervisionService,
                     screening_service: ScreeningService,
                     vac_name_service: VacNameService,
                     ):
    wb = load_workbook(filename=filename)
    service_tabs = ['Child', 'Clinic', 'Parent', 'VacName', 'Dispensary', 'VisitSpecialistControl']

    try:
        ws = wb["Child"]
    except KeyError:
        raise HTTPException(
            status_code=422,
            detail="File has no sheet 'Child'"
        )
    child_gen = create_obj_generator_from_ws(ws=ws, tab_name="Child")
    child = next(child_gen)
    child.kindergarten_num = kindergarten_num

    # Clinic
    try:
        ws = wb["Clinic"]
        clinics_gen = create_obj_generator_from_ws(ws=ws, tab_name="Clinic")
        clinic = next(clinics_gen)
        db_clinic = clinic_service.get_clinic_by_name(name=clinic.name)
        if db_clinic:
            child.clinic_id  = db_clinic.id
        else:
            clinic = clinic_service.add_new_clinic(clinic_data=ClinicCreate(name=clinic.name))
            child.clinic_id = clinic.id
    except KeyError:
        pass

    child.medcard_num = medcard_service.add_new_medcard(child_data=ChildCreate(**child.dict())).medcard_num
    
    
    # Parent
    try:
        ws = wb["Parent"]
        parent_gen = create_obj_generator_from_ws(ws=ws, tab_name="Parent")
        if child.father_id:
            father = next(parent_gen)
            db_father = parent_service.get_paren_by_full_data(parent_data=father)
            if db_father:
                child.father_id = parent_service.add_parent_to_medcard(medcard_num=child.medcard_num,
                                                                    parent_id=db_father.id,
                                                                    parent_type=ParentType.FATHER)
            else:
                father = parent_service.add_new_parent(medcard_num=child.medcard_num, parent_data=ParentCreate(**father.dict(), parent_type='father'))
                child.father_id = father.id

        if child.mother_id:
            mother = next(parent_gen)
            db_mother = parent_service.get_paren_by_full_data(parent_data=mother)
            if db_mother:
                child.mother_id = parent_service.add_parent_to_medcard(medcard_num=child.medcard_num,
                                                                    parent_id=db_mother.id,
                                                                    parent_type=ParentType.MOTHER)
            else:
                mother = parent_service.add_new_parent(medcard_num=child.medcard_num, parent_data=ParentCreate(**mother.dict(), parent_type='mother'))
                child.mother_id = mother.id
    except KeyError:
        pass

    # Dispensary
    try:
        visit_ws = wb["VisitSpecialistControl"]
        visit_s_control_gen = create_obj_generator_from_ws(ws=visit_ws, tab_name="VisitSpecialistControl")
        visit_s_control_list = list(visit_s_control_gen)
    except KeyError:
        pass

    try:
        ws = wb["Dispensary"]
        dispensaryes_gen = create_obj_generator_from_ws(ws=ws, tab_name="Dispensary")
        for dispensary in dispensaryes_gen:
            old_dispensary_id = dispensary.id
            dispensary.medcard_num = child.medcard_num
            dispensary = dispensary_service.add_new_dispensary(dispensary_data=DispensaryCreate(**dispensary.dict()))
            
            for visit_s_control in visit_s_control_list[:]:
                if visit_s_control.dispensary_id == old_dispensary_id:
                    visit_s_control.dispensary_id = dispensary.id
                    visit_s_control_service.add_new_visit_specialist_control(visit_specialist_control_data=VisitSpecialistControlCreate(**visit_s_control.dict()))
                    visit_s_control_list.remove(visit_s_control)
    except KeyError:
        pass


    services = {
        "Child": medcard_service,
        "Allergy": allergy_service,
        "Clinic": clinic_service,
        "Deworming": deworming_service,
        "Dispensary": dispensary_service,
        "ExtraClass": extra_class_service,
        "GammaGlobulinInjection": gg_injection_service,
        "Hospitalization": hospitalization_service,
        "MantouxTest": mantoux_test_service,
        "MedicalCertificate": medical_certificate_service,
        "MedicalExamination": medical_examination_service,
        "OngoingMedicalSupervision": ongoing_medical_supervision_service,
        "OralSanation": oral_sanation_service,
        "Parent": parent_service,
        "PastIllness": past_illness_service,
        "PrevaccinationCheckup": prevaccination_checkup_service,
        "Screening": screening_service,
        "SpaTreatment": spa_treatment_service,
        "TuberculosisVaccination": tub_vac_service,
        "VacName": vac_name_service,
        "Vaccination": vaccination_service,
        "VisitSpecialistControl": visit_s_control_service
    }

    create_model = {
        "Child": ChildCreate,
        "Allergy": AllergyCreate,
        "Clinic": ClinicCreate,
        "Deworming": DewormingCreate,
        "Dispensary": DispensaryCreate,
        "ExtraClass": ExtraClassCreate,
        "GammaGlobulinInjection": GammaGlobulinInjectionCreate,
        "Hospitalization": HospitalizationCreate,
        "MantouxTest": MantouxTestCreate,
        "MedicalCertificate": MedicalCertificateCreate,
        "MedicalExamination": MedicalExaminationCreate,
        "OngoingMedicalSupervision": OngoingMedicalSupervisionCreate,
        "OralSanation": OralSanationCreate,
        "Parent": ParentCreate,
        "PastIllness": PastIllnessCreate,
        "PrevaccinationCheckup": PrevaccinationCheckupCreate,
        "Screening": ScreeningCreate,
        "SpaTreatment": SpaTreatmentCreate,
        "TuberculosisVaccination": TuberculosisVaccinationCreate,
        "VacName": VacNameCreate,
        "Vaccination": VaccinationCreate,
        "VisitSpecialistControl": VisitSpecialistControlCreate
    }

    for tab_name in wb.sheetnames:
        if tab_name not in service_tabs:
            ws = wb[tab_name]

            service = services[tab_name]
            method_suffix = camel_to_snake(tab_name)
            add_method_name = f"add_new_{method_suffix}"
            table = create_obj_generator_from_ws(ws=ws, tab_name=tab_name)

            if hasattr(service, add_method_name):
                add_method = getattr(service, add_method_name)
                try:
                    for table_row in table:
                        table_row.medcard_num = child.medcard_num
                        add_method(create_model[tab_name](**table_row.dict()))
                except AttributeError:
                    print(f"Модель с именем {tab_name} не найдена")
                

                

            