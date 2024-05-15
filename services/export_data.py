from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from models.base_table_name import BaseTableNameModel
from tables import Child


def create_file_with_child_tab(filename: str, child_data: Child) -> None:
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = BaseTableNameModel.Child.value

    headers = [column.name for column in child_data.__table__.columns]
    ws.append(headers)

    row = [getattr(child_data, column.name) for column in child_data.__table__.columns]
    ws.append(row)
    
    wb.save(filename)

def append_data_in_xlsx_file(data: list[BaseModel], tab_name: BaseTableNameModel, filename: str):
    if data:
      wb = load_workbook(filename)
      ws = wb.create_sheet(title=tab_name.value)

      headers = [column.name for column in data[0].__table__.columns]
      ws.append(headers)

      for item in data:
          if item:
            row = [getattr(item, column.name) for column in data[0].__table__.columns]
            ws.append(row)
      wb.save(filename)
